#!/usr/bin/env python3
# /// script
# requires-python = ">=3.9"
# dependencies = [
#     "flask",
#     "openpyxl",
#     "watchdog",
#     "filelock",
# ]
# ///
"""
Live spreadsheet <-> HTML sync server.

Treats a spreadsheet file (.xlsx/.xlsm or .csv) as the backend database and
serves its contents as JSON over a small HTTP API. The HTML frontend polls
for changes and can write cells / rows back. A filesystem watcher also
detects edits made directly in Excel/LibreOffice (or any other program) and
reflects them on the next poll.

Writes do not hit the disk one at a time. Saving a workbook costs seconds
on a large one, so an edit lands in a workbook held open in memory and a
single save covers a burst of them (see _schedule_flush). Anything that
reads the file afterwards flushes first, so the deferral is invisible to
the browser; a save that fails is reported on the next poll rather than on
the request that made the edit.

Known scaling limits, measured: a change of any size invalidates the whole
read cache, so one edit costs a full re-parse and a full payload. See
dev/sync-architecture-review.md before optimising further.

This script is a *template* meant to be copied into a project and adapted:
adjust SHEET file path via --file, and tweak read_workbook/write_cell if the
target spreadsheet needs custom handling (merged cells, multiple header
rows, etc). The header row is auto-detected per sheet (first "wide" row),
so metadata rows commonly found above the real table in exported reports
(company name, report date, etc.) don't get misread as column headers.

Usage (uv resolves the dependencies above automatically, and downloads a
managed Python interpreter if the machine has none):
    uv run sync_server.py --file "budget.xlsx" --ui "app.html" --port 5000
    uv run sync_server.py --file "data.csv" --port 5000        # fallback UI
    uv run --with xlrd sync_server.py --file "legacy_report.xls" --port 5000

Or with a plain Python install:
    pip install flask openpyxl watchdog filelock
    python sync_server.py --file "budget.xlsx" --ui "app.html" --port 5000

The server itself is UI-agnostic: it serves the composed HTML given via
--ui (plus its sibling js/css files) and exposes the spreadsheet as a JSON
API. What the page looks like is decided per-spreadsheet by whoever
composes it from the building blocks in assets/ (see SKILL.md).

Dependencies: flask, openpyxl, watchdog, filelock (declared in the inline
script metadata above, so `uv run` needs no separate install step).
openpyxl is only exercised for .xlsx/.xlsm files, not plain .csv. xlrd is
only needed for legacy .xls files: `uv run --with xlrd` or `pip install
xlrd`. Keep the metadata block and this list in sync with the imports.

Legacy .xls files: openpyxl can't read or write the old binary Excel format
at all, and there's no actively-maintained pure-Python writer for it either.
Rather than fail outright or silently bolt on a fragile legacy writer, this
script converts a .xls to a sibling .xlsx once on startup (via xlrd) and
operates on that from then on. The original .xls is left untouched. The
conversion prints to the console rather than happening quietly, because it
changes which file is now the live source of truth.
"""

import argparse
import contextlib
import csv
import re
import threading
import time
from pathlib import Path

from filelock import FileLock
from flask import Flask, jsonify, request, send_from_directory
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

app = Flask(__name__)

STATE = {
    # --- what we're serving -------------------------------------------
    "path": None,           # the spreadsheet, after any .xls conversion
    "lock_path": None,      # cross-process lock guarding writes
    "ui_dir": None,         # directory the composed HTML lives in
    "ui_file": None,
    "recalc": False,        # --recalc, see _formula_solution_for

    # --- read side ----------------------------------------------------
    # version is the change counter the browser polls; cache is the parsed
    # workbook it was built from, reused until version moves.
    "version": 0,
    "cache": None,
    "cache_version": -1,

    # --- write side ---------------------------------------------------
    # One workbook is held open so an edit costs a save rather than a
    # parse plus a save. wb_path is where a flush writes: the file that
    # workbook came from, which is not always STATE["path"] when these
    # functions are used directly. header_cache saves a per-write scan.
    "wb": None,
    "wb_path": None,
    "header_cache": {},

    # --- deferred saving ----------------------------------------------
    # Saving is O(workbook), so edits land in memory and one save covers a
    # burst of them. flush_error carries a failure back to the browser,
    # which has long since had its response by the time the save runs.
    "dirty": False,
    "flush_timer": None,
    "flush_error": None,

    # --- telling our own writes apart from Excel's --------------------
    # The watcher fires during a multi-second save, when no timestamp or
    # mtime comparison can identify the write as ours, so writes are
    # bracketed instead (see _self_write). The other two are backstops.
    "writing": 0,
    "last_self_write": 0.0,
    "self_mtime": None,

    # --- rows we created that are still empty -------------------------
    # openpyxl writes an empty string as a blank cell, making a new row
    # indistinguishable from a template's reserved ones, so read_workbook
    # would hide it the moment it was added. sheet name -> row indices.
    "appended": {},
}

# Flask sorts dict keys when serializing by default, which reorders the
# sheets alphabetically. Sheet order in a workbook carries meaning (a
# workbook of month sheets is in calendar order, not Août/Avril/Décembre),
# so preserve the order read_workbook built.
try:
    app.json.sort_keys = False           # Flask >= 2.2
except AttributeError:                   # older Flask
    app.config["JSON_SORT_KEYS"] = False


# ---------------------------------------------------------------------------
# Spreadsheet I/O (xlsx via openpyxl, csv via stdlib csv)
# ---------------------------------------------------------------------------

def _is_xlsx(path):
    return path.suffix.lower() in (".xlsx", ".xlsm")


def _is_legacy_xls(path):
    return path.suffix.lower() == ".xls"


def convert_xls_to_xlsx(path):
    """Read a legacy .xls (via xlrd) and write a sibling .xlsx (via openpyxl)
    with the same sheets/data, preserving dates. Returns the new path.
    Skips the conversion if that .xlsx already exists and is newer than the
    .xls, so re-running the server doesn't reconvert (and discard live
    edits already made to the .xlsx) every time.
    """
    import xlrd
    import openpyxl

    out_path = path.with_suffix(".xlsx")
    if out_path.exists() and out_path.stat().st_mtime >= path.stat().st_mtime:
        return out_path

    xls_wb = xlrd.open_workbook(path)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    for name in xls_wb.sheet_names():
        sh = xls_wb.sheet_by_name(name)
        ws = out_wb.create_sheet(name)
        for r in range(sh.nrows):
            row_values = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row_values.append(xlrd.xldate_as_datetime(cell.value, xls_wb.datemode))
                else:
                    row_values.append(cell.value)
            ws.append(row_values)
    out_wb.save(out_path)
    return out_path


def _used_bounds(ws):
    """Return (max_row, max_col) covering only cells that actually hold a
    value.

    `ws.max_row`/`ws.max_column` report the sheet's *declared* dimension,
    which Excel frequently writes far larger than the real data. A template
    workbook where a whole column band was formatted, or where a filler row
    pins the table's shape, commonly declares the full 16,384 columns and
    thousands of rows while holding a handful of real ones. Three things
    break when that number is trusted:

      - _detect_header_row's "wide row" threshold becomes half of 16,384, so
        no row ever qualifies and every sheet falls back to header row 1.
      - iter_rows walks 16,384 columns per row, turning a fast parse into a
        30-second one and inventing thousands of `colN` placeholder headers.
      - append_row lands the new row below the phantom rows instead of after
        the real data, outside whatever ranges the workbook's formulas cover.

    openpyxl stores cells sparsely, so scanning the populated ones is cheap
    and gives the true extent. Styled-but-empty cells are present in that
    store with a None value, which is exactly what the value check filters
    out. Falls back to the declared dimension if the internal store isn't
    available (a future openpyxl, or read-only mode).
    """
    cells = getattr(ws, "_cells", None)
    if not cells:
        return (ws.max_row or 1), (ws.max_column or 1)
    max_row = max_col = 0
    for (r, c), cell in cells.items():
        if cell.value is not None:
            if r > max_row:
                max_row = r
            if c > max_col:
                max_col = c
    return (max_row or 1), (max_col or 1)


def _used_bounds_from_rows(*grids):
    """Same idea as _used_bounds, for rows already streamed as plain value
    tuples. read_only mode has no cell store to inspect, and its reported
    dimensions are no more trustworthy than the normal mode's, so the extent
    is measured from the values themselves.
    """
    max_row = max_col = 0
    for grid in grids:
        for r, row in enumerate(grid, start=1):
            for c, value in enumerate(row, start=1):
                if value is not None:
                    if r > max_row:
                        max_row = r
                    if c > max_col:
                        max_col = c
    return (max_row or 1), (max_col or 1)


def _headers_from_row(header_values):
    """Column names for a header row.

    Trailing columns with no header are dropped: a table is as wide as its
    header row, and Excel parks data-validation source lists in far-off
    columns of the same sheet, which the used-range scan otherwise sweeps in
    along with the empty padding between. Unnamed columns *inside* the table
    keep a colN placeholder so positions still line up with the cells.

    The read path and the write path both name columns through here. If
    they derived names separately they could disagree about which row is
    the header, and a write would then fail to find its column.
    """
    cells = list(header_values)
    last_named = max((i for i, v in enumerate(cells) if v is not None), default=-1)
    if last_named >= 0:
        cells = cells[:last_named + 1]
    return [v if v is not None else f"col{i}" for i, v in enumerate(cells)]


def _detect_header_row(rows_iter, max_col, scan_limit=30):
    """Find the row that looks like the table's header.

    Exported reports often prepend narrow "label, value" metadata rows
    (Company, Report Date) before the real table, so row 1 is not a safe
    assumption. The header is taken to be the first row, within the first
    `scan_limit`, populated across roughly half or more of the columns.

    When no row is "wide" enough to look like a header (a two-column
    reference list on a six-column sheet, say), fall back to the first row
    that holds anything at all rather than to row 1. Sheets often start with
    a blank spacer row, and calling that the header names every column
    `colN` and pushes the real labels down into the data.
    """
    threshold = max(3, max_col * 0.5)
    for i, row in enumerate(rows_iter[:scan_limit], start=1):
        nonnull = sum(1 for v in row if v is not None)
        if nonnull >= threshold:
            return i
    for i, row in enumerate(rows_iter[:scan_limit], start=1):
        if any(v is not None for v in row):
            return i
    return 1


def _extract_meta(rows_before_header):
    """Collect the narrow "label, value" rows above the header into a dict,
    so context like a company name or report date is surfaced rather than
    dropped. Exposed to the frontend as sheet.meta.
    """
    meta = {}
    for row in rows_before_header:
        vals = [v for v in row if v is not None]
        if len(vals) == 2:
            meta[str(vals[0])] = vals[1] if not hasattr(vals[1], "isoformat") else vals[1].isoformat()
    return meta




def read_workbook(path, recalc=False, keep_rows=None):
    """Return {sheet_name: {"headers": [...], "rows": [ {col: value}, ... ],
    "formulas": {(row_idx, col): "=..."}, "meta": {...}, "header_row": N }}.
    The header row is auto-detected per sheet rather than assumed to be
    row 1. See _detect_header_row.

    If `recalc` is True, formula cells that have no cached value yet are
    filled in using the `formulas` package, as a display-only enrichment of
    the JSON response. See _formula_solution_for for why this never touches
    the file on disk. A cell most often lacks a cached value when the
    workbook was built with openpyxl and never saved by real Excel, so Excel
    never wrote a result.
    """
    path = Path(path)
    if _is_xlsx(path):
        import openpyxl

        # read_only streams the sheet XML instead of building a full cell
        # model. On a workbook whose sheets carry thousands of pre-formatted
        # rows this is the difference between ~15s and ~0.05s per load, and
        # this function only ever reads. Writes go through
        # _writable_workbook, which needs the normal (editable) mode.
        wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True)
        wb_formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
        solution = _formula_solution_for(path) if recalc else None
        sheets = {}
        try:
            for name in wb_values.sheetnames:
                value_grid = list(wb_values[name].iter_rows(values_only=True))
                formula_grid = list(wb_formulas[name].iter_rows(values_only=True))
                # A formula cell with no cached result is None in the value
                # pass but "=..." in the formula pass, so the real extent is
                # the union of both.
                used_row, used_col = _used_bounds_from_rows(value_grid, formula_grid)
                # A just-appended empty row sits past the last cell holding a
                # value, so the bounds above exclude it and the loop below
                # would never see it at all. Stretch the range to cover it.
                keep = (keep_rows or {}).get(name, ())
                if keep:
                    used_row = max(used_row, max(keep))
                if not value_grid:
                    sheets[name] = {"headers": [], "rows": [], "formulas": {}, "meta": {}, "header_row": 1}
                    continue
                rows_iter = [row[:used_col] for row in value_grid[:used_row]]
                while len(rows_iter) < used_row:      # kept rows past the last stored cell
                    rows_iter.append(tuple([None] * used_col))
                header_row_idx = _detect_header_row(rows_iter, used_col)
                meta = _extract_meta(rows_iter[: header_row_idx - 1])
                headers = _headers_from_row(rows_iter[header_row_idx - 1])
                rows_iter = [row[:len(headers)] for row in rows_iter]
                rows = []
                formula_cells = {}
                for r_idx, row in enumerate(rows_iter[header_row_idx:], start=header_row_idx + 1):
                    f_row = formula_grid[r_idx - 1] if r_idx - 1 < len(formula_grid) else ()
                    # A row with no values isn't a record. Templates reserve
                    # long blocks of pre-formatted rows, often with a formula
                    # already filled down each one; those render blank and
                    # would otherwise flood the grid with hundreds of empty
                    # rows. The test is deliberately on values only, not
                    # formulas: a formula row that has produced a result has
                    # a value and stays. Cells written by "+ Add row" hold an
                    # empty string, which openpyxl writes as an empty cell,
                    # so keep_rows carries the ones this server just added.
                    if all(v is None for v in row) and r_idx not in keep:
                        continue
                    record = {"_row": r_idx}
                    for c_idx, value in enumerate(row):
                        header = headers[c_idx]
                        record[header] = value
                        f_value = f_row[c_idx] if c_idx < len(f_row) else None
                        if isinstance(f_value, str) and f_value.startswith("="):
                            formula_cells[f"{r_idx}:{header}"] = f_value
                            if value is None and solution is not None:
                                record[header] = _formula_solution_lookup(solution, name, c_idx + 1, r_idx)
                    rows.append(record)
                sheets[name] = {"headers": headers, "rows": rows, "formulas": formula_cells, "meta": meta, "header_row": header_row_idx}
        finally:
            wb_values.close()
            wb_formulas.close()
        return sheets
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            rows = []
            for r_idx, record in enumerate(reader, start=2):
                record["_row"] = r_idx
                rows.append(record)
        return {path.stem: {"headers": headers, "rows": rows, "formulas": {}, "meta": {}, "header_row": 1}}


def _coerce_value(value):
    """Turn a browser-supplied string into the type Excel should store.

    A contentEditable cell and a text input both hand back strings, so
    typing 125.5 into a money column used to store the *text* "125.5".
    Excel then leaves it out of SUM, and the workbook's own totals quietly
    stop matching what the page shows.

    Deliberately conservative: only unambiguous forms convert.
      - "" clears the cell (openpyxl stores an empty string as blank anyway)
      - "125", "-3"      -> int
      - "125.5", "-0.5"  -> float
      - "007", "+33..."  -> left as text; leading zeros and plus signs mean
        product codes and phone numbers, not quantities
      - "125,50"         -> left as text. A comma is a decimal point in
        French and a thousands separator in English, and the server can't
        tell which. A page that knows its locale should send a real JSON
        number instead of a string, which arrives here already typed.
    """
    if not isinstance(value, str):
        return value          # JSON numbers and booleans arrive already typed
    s = value.strip()
    if s == "":
        return None
    if re.fullmatch(r"-?\d+", s) and not re.fullmatch(r"-?0\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d*\.\d+", s) and not re.fullmatch(r"-?0\d+\.\d+", s):
        return float(s)
    return value


def _xlsx_headers(ws):
    """Column names of the sheet's header row. See _xlsx_header_info."""
    return _xlsx_header_info(ws)[0]


def _xlsx_header_info(ws):
    """Locate the header row and name its columns, using exactly the same
    logic read_workbook uses, so a write lands in the column the browser
    thinks it is editing. Cached per sheet: the scan is not free and the
    answer only changes when the workbook is reloaded or rows move.
    """
    cached = STATE["header_cache"].get(ws.title)
    if cached is not None:
        return cached
    used_row, used_col = _used_bounds(ws)
    grid = [tuple(c.value for c in row)
            for row in ws.iter_rows(max_row=used_row, max_col=used_col)]
    header_row_idx = _detect_header_row(grid, used_col)
    result = (_headers_from_row(grid[header_row_idx - 1]), header_row_idx)
    STATE["header_cache"][ws.title] = result
    return result


def _next_free_row(ws, headers, header_row_idx, used_row):
    """First row at or after the header that holds no data yet.

    "Append" means different things on the two shapes of sheet this serves.
    On an ordinary list it's the row after the last one. On a template that
    has a formula filled down a long reserved block, the row after the block
    sits outside whatever ranges the workbook's own totals sum over, so a
    sale written there never reaches the dashboard. Looking for the first
    row whose non-formula cells are all empty handles both: on a plain list
    that is the row after the data, and on a template it is the next free
    slot inside the block.
    """
    for r in range(header_row_idx + 1, used_row + 2):
        free = True
        for col_idx in range(1, len(headers) + 1):
            value = ws.cell(row=r, column=col_idx).value
            if isinstance(value, str) and value.startswith("="):
                continue  # a formula is scaffolding, not data
            if value is not None and value != "":
                free = False
                break
        if free:
            return r
    return used_row + 1


def _writable_workbook(path):
    """Return a writable workbook for `path`, kept in memory between writes.

    Parsing a large workbook costs seconds, and the naive path pays it on
    every single edit: load, change one cell, save. On a workbook with
    thousands of styled rows that is tens of seconds per keystroke-sized
    change, which is not a live app. Holding one workbook open makes a write
    cost only the save.

    The cache is dropped whenever the file changes underneath us, so an edit
    made in Excel is never overwritten by a stale in-memory copy. See
    _ChangeHandler and _invalidate_workbook.
    """
    import openpyxl

    if STATE["wb"] is None:
        STATE["wb"] = openpyxl.load_workbook(path, data_only=False)
        # Remember where it came from. A deferred flush must write back to
        # the workbook's own file, not to whatever STATE["path"] happens to
        # hold: they are the same when the server runs, and are not when
        # these functions are used directly, which this template invites.
        STATE["wb_path"] = str(path)
    return STATE["wb"]


def _invalidate_workbook():
    """Drop the cached workbook.

    Pending edits live only in that object, so land them first rather than
    discarding them. The watcher already refuses to invalidate while edits
    are in flight; this covers every other caller, because silently losing
    a change the user already made is the worst outcome available.
    """
    if STATE["dirty"] and STATE["wb"] is not None:
        try:
            _flush()
        except Exception:
            pass
    STATE["wb"] = None
    STATE["wb_path"] = None
    STATE["header_cache"].clear()
    # Whoever edited the file may have filled or removed the blank rows we
    # were keeping visible; re-reading from disk is the honest state.
    STATE["appended"].clear()


def _warm_workbook(path):
    """Load the writable workbook in the background at startup.

    The read path is fast (read_only mode), but the write path needs the
    editable model, which on a heavily styled workbook costs ~20s to build.
    Paying that on the user's first edit makes the first edit feel broken.
    Doing it during startup instead means the cost lands while nobody is
    waiting.

    The result is only adopted if nothing changed while it loaded: any write
    or external edit bumps the version, and adopting a workbook parsed
    before that point could write stale content back over it.
    """
    import openpyxl

    version_at_start = STATE["version"]
    try:
        wb = openpyxl.load_workbook(path, data_only=False)
    except Exception:
        return  # the first real write will surface the problem properly
    if STATE["wb"] is None and STATE["version"] == version_at_start:
        STATE["wb"] = wb


def write_cell(path, sheet, row_idx, column, value):
    path = Path(path)
    if _is_xlsx(path):
        wb = _writable_workbook(path)
        ws = wb[sheet]
        headers = _xlsx_headers(ws)
        col_idx = headers.index(column) + 1
        ws.cell(row=row_idx, column=col_idx, value=_coerce_value(value))
        _schedule_flush()
    else:
        rows = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            for r_idx, record in enumerate(reader, start=2):
                if r_idx == row_idx:
                    record[column] = value
                rows.append(record)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)


def append_row(path, sheet, values):
    path = Path(path)
    if _is_xlsx(path):
        wb = _writable_workbook(path)
        ws = wb[sheet]
        headers, header_row_idx = _xlsx_header_info(ws)
        # ws.append() targets the declared dimension's next row, which on a
        # template sheet is below thousands of reserved rows and outside the
        # ranges the workbook's own formulas sum over. Find the first row
        # that actually has no data instead.
        used_row, _ = _used_bounds(ws)
        target = _next_free_row(ws, headers, header_row_idx, used_row)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=target, column=col_idx)
            # A template can carry a formula in the row being appended to
            # (a running total, a computed column). Leave those alone; every
            # other column takes the submitted value, or an empty string so
            # the new row is visible while the user fills it in.
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = _coerce_value(values.get(header, ""))
        _schedule_flush()
        # See STATE["appended"]: an all-empty append is invisible on disk.
        if not any(str(v).strip() for v in values.values()):
            STATE["appended"].setdefault(sheet, set()).add(target)
        return target
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            existing = sum(1 for _ in reader)
        with open(path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writerow(values)
        return existing + 2  # header is row 1, so the first data row is 2


def delete_row(path, sheet, row_idx):
    path = Path(path)
    if _is_xlsx(path):
        wb = _writable_workbook(path)
        ws = wb[sheet]
        ws.delete_rows(row_idx)
        # Deleting shifts everything below it, so any cached header position
        # for this sheet is no longer trustworthy.
        STATE["header_cache"].pop(sheet, None)
        _schedule_flush()
    else:
        rows = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            for r_idx, record in enumerate(reader, start=2):
                if r_idx != row_idx:
                    rows.append(record)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)


_FORMULA_SOLUTION_KEY_RE = re.compile(r"^'\[[^\]]*\](?P<sheet>.*)'!(?P<ref>[A-Z]+\d+)$")


def _formula_solution_for(path):
    """Evaluate every formula in the workbook using the optional `formulas`
    package and return its raw solution dict (cell-ref -> computed value),
    or None if the package isn't installed or evaluation fails outright
    (unsupported functions, circular refs, and so on). Individual
    unsupported cells are handled per-lookup in _formula_solution_lookup
    instead.

    Deliberately read-only. An earlier version recalculated by asking
    `formulas` to rewrite the xlsx via `ExcelModel.write`, which rebuilds
    the entire workbook from only the cells it touched during evaluation.
    On a real file that silently dropped data and formatting outside the
    formula dependency graph, and permanently replaced formula cells with
    plain numbers. openpyxl can't write a cached value alongside a formula
    the way Excel does, so there is no non-destructive way to persist this
    to the file at all. Enriching only the JSON the browser sees avoids the
    risk, at the cost of the recalculated value being invisible if the user
    opens the raw file in Excel.
    """
    try:
        import numpy as np
        import formulas

        xl = formulas.ExcelModel().loads(str(path)).finish()
        raw = xl.calculate()
        # Index once by (SHEET, CELLREF) rather than re-scanning the whole
        # solution per formula cell. A real workbook can have thousands.
        indexed = {}
        for key, ranges in raw.items():
            m = _FORMULA_SOLUTION_KEY_RE.match(key)
            if not m:
                continue
            try:
                value = np.ravel(ranges.value)[0]
                indexed[(m.group("sheet").strip().upper(), m.group("ref"))] = value.item() if hasattr(value, "item") else value
            except Exception:
                continue
        return indexed
    except Exception:
        return None


def _formula_solution_lookup(solution, sheet_name, col_idx, row_idx):
    import openpyxl.utils

    cell_ref = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
    return solution.get((sheet_name.strip().upper(), cell_ref))


# ---------------------------------------------------------------------------
# Filesystem watcher: detect edits made outside our own writes (e.g. a human
# editing the file directly in Excel) and bump the version so pollers refresh
# ---------------------------------------------------------------------------

FLUSH_DELAY = 0.5   # seconds of quiet before a burst of edits is written


def _lock():
    """The cross-process file lock, or nothing when there's no path set.

    This module is meant to be copied and adapted, and its read/write
    functions are usable on their own without running the server. In that
    case there is no lock path and no second writer to guard against.
    """
    if not STATE["lock_path"]:
        return contextlib.nullcontext()
    return FileLock(STATE["lock_path"])


def _schedule_flush():
    """Mark the cached workbook dirty and (re)arm the save timer.

    Each edit resets the timer, so typing across several cells costs one
    save instead of one per cell. The version counter is deliberately not
    bumped here: the browser already shows what the user typed, and bumping
    early would make it refetch a file that doesn't have the change yet.
    """
    STATE["dirty"] = True
    timer = STATE["flush_timer"]
    if timer is not None:
        timer.cancel()
    timer = threading.Timer(FLUSH_DELAY, _flush)
    timer.daemon = True
    STATE["flush_timer"] = timer
    timer.start()


def _flush():
    """Write the cached workbook to disk, if anything is pending.

    Failures (the file open in Excel, most often) are recorded and reported
    on the next /api/version poll, because the HTTP request that made the
    edit has long since returned by the time this runs.
    """
    with _lock():
        if not STATE["dirty"] or STATE["wb"] is None:
            return
        target = STATE["wb_path"] or STATE["path"]
        if not target:
            return
        try:
            with _self_write():
                STATE["wb"].save(target)
            STATE["dirty"] = False
            STATE["flush_error"] = None
            STATE["version"] += 1
        except PermissionError:
            STATE["flush_error"] = ("The spreadsheet is locked by another program. "
                                    "Close it in Excel and your changes will be saved.")
        except Exception as exc:
            STATE["flush_error"] = f"Could not save: {type(exc).__name__}: {exc}"


def flush_now():
    """Force any pending edits to disk and report whether it worked."""
    timer = STATE["flush_timer"]
    if timer is not None:
        timer.cancel()
    _flush()
    return STATE["flush_error"]


class _self_write:
    """Bracket a save so the watcher doesn't mistake it for someone else's.

    The watcher fires mid-save, when the file is neither in its old state
    nor its final one. Without a bracket it judges the change external and
    drops the cached workbook, so every edit pays a full reload.
    """

    def __enter__(self):
        STATE["writing"] += 1
        STATE["last_self_write"] = time.time()
        return self

    def __exit__(self, *exc):
        _mark_self_write()
        STATE["writing"] -= 1
        return False


def _mark_self_write():
    """Record the state our own save left the file in, so the watcher can
    recognise it. Called after the write, not before: a save on a large
    workbook takes seconds, and a timestamp taken beforehand has already
    aged out by the time the filesystem event arrives.
    """
    STATE["last_self_write"] = time.time()
    try:
        STATE["self_mtime"] = Path(STATE["wb_path"] or STATE["path"]).stat().st_mtime
    except (OSError, TypeError):
        STATE["self_mtime"] = None


class _ChangeHandler(FileSystemEventHandler):
    def on_modified(self, event):
        if event.src_path and Path(event.src_path).resolve() == Path(STATE["path"]).resolve():
            # Ignore the event our own write_cell/append_row just caused.
            # Match on the mtime our save produced, with the timestamp as a
            # backstop for filesystems with coarse mtime resolution.
            try:
                current = Path(event.src_path).stat().st_mtime
            except OSError:
                current = None
            ours = STATE["writing"] > 0 or (
                STATE["self_mtime"] is not None and current is not None
                and abs(current - STATE["self_mtime"]) < 0.001)
            if not ours and STATE["dirty"]:
                return   # our own unsaved edits are still in flight
            if not ours and time.time() - STATE["last_self_write"] > 0.75:
                STATE["version"] += 1
                # Somebody else changed the file, so the workbook held open
                # for writes is now stale. Dropping it forces the next write
                # to re-read from disk; saving the stale copy instead would
                # silently revert whatever was just done in Excel.
                _invalidate_workbook()


def start_watcher(path):
    observer = Observer()
    observer.schedule(_ChangeHandler(), str(Path(path).parent), recursive=False)
    observer.daemon = True
    observer.start()
    return observer


# ---------------------------------------------------------------------------
# HTTP API
# ---------------------------------------------------------------------------

def _get_sheets():
    """Re-parsing a multi-thousand-row workbook (twice, for values + formulas)
    takes over a second on real-world files. Since the frontend polls
    frequently but the file usually hasn't changed between polls, cache the
    parsed result and only re-read from disk when the version counter moved
    (from our own write or the file watcher noticing an external edit).
    """
    # Pending edits live only in the cached workbook, so land them before
    # re-reading from disk or the browser would be shown its own edit
    # reverted.
    if STATE["dirty"]:
        flush_now()
    if STATE["cache"] is None or STATE["cache_version"] != STATE["version"]:
        with _lock():
            STATE["cache"] = read_workbook(STATE["path"], recalc=STATE["recalc"], keep_rows=STATE["appended"])
        STATE["cache_version"] = STATE["version"]
    return STATE["cache"]


def _write_guard(fn):
    """Run a write and turn a refusal into a JSON answer the page can show.

    The common case is Excel holding the file open on Windows, which surfaces
    as PermissionError. Flask's default 500 is an HTML page, so a browser
    that surfaces it verbatim shows the user a stack-trace page instead of
    "close the file in Excel". 423 Locked is the honest status.
    """
    try:
        return fn()
    except PermissionError:
        return jsonify({
            "ok": False,
            "error": "The spreadsheet is locked by another program. "
                     "Close it in Excel (or save and close) and try again.",
        }), 423
    except Exception as exc:
        return jsonify({"ok": False, "error": f"{type(exc).__name__}: {exc}"}), 500


@app.get("/api/version")
def api_version():
    """Cheap endpoint for the frontend to poll frequently without paying the
    cost of re-serializing the whole sheet when nothing has changed.
    """
    body = {"version": STATE["version"]}
    if STATE["flush_error"]:
        # The request that made the edit returned before the save ran, so
        # this poll is the first chance to tell the page it didn't land.
        body["error"] = STATE["flush_error"]
    return jsonify(body)


@app.get("/api/data")
def api_data():
    return jsonify({"version": STATE["version"], "sheets": _get_sheets()})


@app.post("/api/cell")
def api_cell():
    body = request.get_json(force=True)
    def run():
        with _lock():
            write_cell(STATE["path"], body["sheet"], int(body["row"]), body["column"], body["value"])
        return jsonify({"ok": True, "version": STATE["version"], "pending": True})
    return _write_guard(run)


@app.post("/api/rows")
def api_add_row():
    body = request.get_json(force=True)
    def run():
        with _lock():
            row = append_row(STATE["path"], body["sheet"], body["values"])
        # The row index goes back so the page can act on what it just
        # created (open a form on it, scroll to it). Without it a
        # composition can add a row but has no way to find it.
        return jsonify({"ok": True, "version": STATE["version"], "row": row, "pending": True})
    return _write_guard(run)


@app.delete("/api/rows/<sheet>/<int:row_idx>")
def api_delete_row(sheet, row_idx):
    def run():
        with _lock():
            delete_row(STATE["path"], sheet, row_idx)
        return jsonify({"ok": True, "version": STATE["version"], "pending": True})
    return _write_guard(run)


@app.get("/")
def index():
    return send_from_directory(STATE["ui_dir"], STATE["ui_file"])


@app.get("/<path:fname>")
def ui_static(fname):
    """Serve sibling files of the composed UI (sheetsync.js, components.js,
    base.css, images, ...) so a composed app is just a directory of plain
    files next to each other. /api/* routes are registered above and take
    precedence over this catch-all.
    """
    return send_from_directory(STATE["ui_dir"], fname)


# Ports Chrome (and most Chromium browsers) refuse to connect to at all,
# answering ERR_UNSAFE_PORT. A server bound here works perfectly from curl
# and looks completely broken in the browser. Refusing to start says so
# up front rather than leaving it to be discovered.
BLOCKED_BROWSER_PORTS = {
    1, 7, 9, 11, 13, 15, 17, 19, 20, 21, 22, 23, 25, 37, 42, 43, 53, 69, 77,
    79, 87, 95, 101, 102, 103, 104, 109, 110, 111, 113, 115, 117, 119, 123,
    135, 137, 138, 139, 143, 161, 179, 389, 427, 465, 512, 513, 514, 515,
    526, 530, 531, 532, 540, 548, 554, 556, 563, 587, 601, 636, 989, 990,
    993, 995, 1719, 1720, 1723, 2049, 3659, 4045, 4190, 5060, 5061, 6000,
    6566, 6665, 6666, 6667, 6668, 6669, 6679, 6697, 10080,
}


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, help="Path to the .xlsx/.xlsm/.csv/.xls spreadsheet")
    parser.add_argument("--port", type=int, default=5000)
    parser.add_argument("--ui", help="Path to the composed HTML UI to serve at /. Its directory is also served as static files, so keep sheetsync.js/components.js/base.css next to it")
    parser.add_argument("--recalc", action="store_true", help="Fill in formula cells that have no cached value yet using the `formulas` package. Display only, never written to the file (pip install formulas)")
    args = parser.parse_args()

    path = Path(args.file).resolve()
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    if args.ui:
        ui_path = Path(args.ui).resolve()
        if not ui_path.exists():
            raise SystemExit(f"UI file not found: {ui_path}")
    else:
        # Fall back to the bundled example composition (generic grid+search).
        ui_path = Path(__file__).parent.parent / "assets" / "example-app.html"
    STATE["ui_dir"] = str(ui_path.parent)
    STATE["ui_file"] = ui_path.name

    if _is_legacy_xls(path):
        print(f"{path.name} is a legacy .xls file. openpyxl can't read or write that format.")
        new_path = convert_xls_to_xlsx(path)
        print(f"Converted to {new_path.name} (original left untouched). Serving the .xlsx from now on.")
        path = new_path

    STATE["path"] = path
    STATE["lock_path"] = str(path) + ".lock"
    STATE["recalc"] = args.recalc

    if args.port in BLOCKED_BROWSER_PORTS:
        raise SystemExit(
            f"Port {args.port} is on Chrome's blocked list, so the browser will refuse to "
            f"connect (ERR_UNSAFE_PORT) even though the server would run fine. "
            f"Pick another port, for example {args.port + 2}."
        )

    observer = start_watcher(path)
    if _is_xlsx(path):
        threading.Thread(target=_warm_workbook, args=(path,), daemon=True).start()
    print(f"Serving {path.name} live at http://127.0.0.1:{args.port}")
    print("Edit the file in Excel or in the browser. Both sides stay in sync.")
    try:
        app.run(host="127.0.0.1", port=args.port, threaded=True)
    finally:
        # A burst of edits may still be sitting in memory; land them before
        # the process goes away.
        try:
            err = flush_now()
            if err:
                print(f"Warning: unsaved changes could not be written. {err}")
        except Exception as exc:  # pragma: no cover - shutdown best effort
            print(f"Warning: failed to flush pending changes: {exc}")
        observer.stop()


if __name__ == "__main__":
    main()
